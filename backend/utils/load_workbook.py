import json
from openpyxl import load_workbook
from datetime import date, datetime


def excel_to_jsonl_with_formulas(excel_path, jsonl_path):
    wb = load_workbook(excel_path, data_only=False)  # keep formulas

    with open(jsonl_path, "w") as f:
        for sheet in wb.sheetnames:
            ws = wb[sheet]

            # 🔹 Find the first non-empty row (use as header row)
            header_row = None
            headers = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                values = [cell.value for cell in row]
                if any(values):  # row has at least one non-empty cell
                    header_row = row[0].row
                    headers = values
                    break

            if not headers:  # skip if sheet is completely empty
                continue

            # 🔹 Iterate rows AFTER header row
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=header_row + 1), start=header_row + 1
            ):
                row_data = {"sheet": sheet, "row": row_idx, "cells": []}

                for col_idx, cell in enumerate(row, start=1):
                    col_letter = cell.column_letter
                    header = headers[col_idx - 1] if col_idx <= len(headers) else None

                    # Handle values safely
                    val = cell.value
                    if isinstance(val, (datetime, date)):
                        val = val.isoformat()  # "2024-01-01" etc.

                    row_data["cells"].append(
                        {
                            "cell": f"{col_letter}{row_idx}",
                            "header": header,
                            "value": val,
                            "formula": getattr(cell, "formula", None),
                        }
                    )

                f.write(json.dumps(row_data) + "\n")
                